import openpyxl
import sys
import os
import csv
from openpyxl.utils import get_column_letter

p = os.path.dirname(os.path.realpath('f.py'))
p=p+"/"+os.pardir
carrier = sys.argv[1]
group = sys.argv[2]
timeStamp = sys.argv[3]

#Create an excel workbook
wb=openpyxl.Workbook()


reports = p+"/reports/"

#All of Dewey's files are created with this at the end
endStr = "_"+carrier+"_"+group+"_"+timeStamp+".txt"

#Titles of validator sheets
sheetTitles=['File Detail',"Member Counts","Member Missing Data","Verification Table","Members Table","Products Table","Vol Life Table","Coverage Level Test"]
#Titles of text files used to pull data into validator
files=['file_detail','member_counts','member_missing_data','verification_table','members_table','products_table','vol_life_table','coverage_level_report']

ws={}
i=0

def autofitColumns(sheet):
	dims = {}
	#The original validator is autoformatted with variable column widths based on the width of cells. This mimics that.
	for row in sheet.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
	for col, value in dims.items():
		sheet.column_dimensions[col].width = value
hasMassMutual = False;
for fileName in files:

	#Create sheets within workbook
	if i==0:
		ws[0]=wb.active
		ws[0].title = 'File Detail'
	else:
		ws[i]=wb.create_sheet(title=sheetTitles[i])
	f = open(reports+fileName+endStr)
	
	#Apparently this was the only modification from the original text files: adding a title for the first sheet
	if i==0:
		ws[i].append(['File Detail Categories','File Detail'])
	
	#Here's where we extract data from the text files. They are newline delimited and tab delimited.
	k=0
	#This function gets the data from a file and splits it into an array with \n delimiter
	reader = csv.reader(f, delimiter='\n')
	for row in reader:
		#Now we split each row by tabs. 
		row=row[0].split('\t')
		if k==0:
			numColumns=len(row)
		if i==0:
			if len(row)>0 and row[0]=="PayerNameMaxwellID:" and row[1]=="558d94309b22128f141f1bf4":
				hasMassMutual=True;
		ws[i].append(row)
		k+=1
	autofitColumns(ws[i])
	#The original validator has filters enabled by default. This mimics that.
	ws[i].auto_filter.ref = "A1:"+get_column_letter(numColumns)+"1"
	f.close()
	i+=1

#Re-input and then add Pat's worksheet
if hasMassMutual:
	f = open(reports + "massMutualBeneficiaries"+endStr)
	reader = csv.reader(f, delimiter='\n')
	ws=wb.create_sheet(title="MassMutual File Review")
	for row in reader:
		row = row[0].split('\t')
		numColumns=len(row)
		ws.append(row)
	#autofitColumns(ws) #This one seems to work better not auto-fit!
	ws.auto_filter.ref = "A1:"+get_column_letter(numColumns)+"1"
	f.close()

#The workbook is autosaved... no longer have to do it by hand.
wb.save(p+'/excel_reports/'+carrier+"_"+group+"_"+timeStamp+".xlsx")
print ("Created workbook: "+carrier+"_"+group+"_"+timeStamp+".xlsx")